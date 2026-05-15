import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
tier_colors = {'1': 'C6EFCE', '2': 'DDEBF7', '3': 'FFF2CC'}

def header(ws, cols, widths, color):
    ws.append(cols)
    for i, (cell, w) in enumerate(zip(ws[1], widths), 1):
        cell.font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
        cell.fill = PatternFill('solid', start_color=color)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

def row_style(ws, r, even):
    for cell in ws[r]:
        cell.fill = PatternFill('solid', start_color=('EBF3FB' if even else 'FFFFFF'))
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 17

def tier_cell(ws, r, col, val):
    cell = ws.cell(row=r, column=col)
    t = str(val)
    if t in tier_colors:
        cell.fill = PatternFill('solid', start_color=tier_colors[t])
        cell.font = Font(bold=True, name='Calibri')
        cell.alignment = Alignment(horizontal='center', vertical='center')

# ── SHEET 1: COMPANIES ────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Companies'
ws1.freeze_panes = 'A2'
header(ws1,
    ['Company','Careers URL','LinkedIn Slug','Tier','Priority','Active','Cities','Sector','Notes'],
    [28,48,26,8,10,8,36,22,30], '1F4E79')
ws1.auto_filter.ref = 'A1:I1'

companies = [
# TIER 1
('EXL Service','https://www.exlservice.com/careers','exl',1,3,'TRUE','Mumbai, Pune, Bangalore, Hyderabad','Analytics BPO','Largest BFSI analytics BPO; hires 100s of analysts/yr'),
('Genpact','https://www.genpact.com/careers','genpact',1,3,'TRUE','Hyderabad, Bangalore, Pune','Analytics BPO','Massive BFSI analytics division; fresher-friendly'),
('WNS Global Services','https://www.wns.com/careers','wns-global-services',1,3,'TRUE','Mumbai, Pune, Bangalore','Analytics BPO','Finance analytics BPO; structured fresher intake'),
('Accenture','https://www.accenture.com/in-en/careers','accenture',1,3,'TRUE','Mumbai, Pune, Bangalore, Hyderabad','Consulting','Bulk hiring; massive BFSI analytics practice'),
('Capgemini','https://www.capgemini.com/in-en/careers/','capgemini',1,3,'TRUE','Mumbai, Pune, Bangalore, Hyderabad','Consulting','Strong BFSI analytics division'),
('Deloitte','https://jobs2.deloitte.com/in/en','deloitte',1,3,'TRUE','Mumbai, Pune, Bangalore, Hyderabad','Big 4','Hires freshers in analytics + risk consulting'),
('PwC','https://www.pwc.in/careers.html','pwc',1,3,'TRUE','Mumbai, Pune, Bangalore, Hyderabad','Big 4','Analytics COE; active fresher programs'),
('EY','https://careers.ey.com/','ernst-young',1,3,'TRUE','Mumbai, Pune, Bangalore, Hyderabad','Big 4','Data analytics practice; hires freshers'),
('KPMG','https://home.kpmg/in/en/home/careers.html','kpmg-india',1,3,'TRUE','Mumbai, Pune, Bangalore, Hyderabad','Big 4','Analytics + risk consulting fresher intake'),
('HDFC Bank','https://www.hdfcbank.com/content/bbp/repositories/723fb80a-2dde-42a3-9793-7ae1be57c87f/?folderPath=/footer/Careers/','hdfc-bank',1,3,'TRUE','Mumbai, Pune, Bangalore, Hyderabad','Banking','Largest private bank; huge data team'),
('ICICI Bank','https://www.icicicareers.com/','icici-bank',1,3,'TRUE','Mumbai, Pune, Bangalore, Hyderabad','Banking','Strong analytics COE; hires freshers'),
('Axis Bank','https://www.axisbank.com/careers','axis-bank',1,3,'TRUE','Mumbai, Pune, Bangalore','Banking','Growing data team; regular fresher intake'),
('JPMorgan Chase','https://careers.jpmorgan.com/','jpmorgan-chase',1,3,'TRUE','Mumbai, Bangalore, Hyderabad','Investment Bank GCC','Analyst programs for freshers; large GCC'),
('Citi','https://jobs.citi.com/','citi',1,3,'TRUE','Mumbai, Pune, Bangalore','Banking GCC','Analyst programs; strong GCC data team'),
('Wells Fargo','https://www.wellsfargojobs.com/','wells-fargo',1,3,'TRUE','Hyderabad, Bangalore','Banking GCC','Large GCC; hires data analysts in bulk'),
('Bank of America','https://careers.bankofamerica.com/','bank-of-america',1,3,'TRUE','Mumbai, Hyderabad','Banking GCC','Solid GCC with analytics division'),
('Fractal Analytics','https://fractal.ai/careers/','fractal-analytics',1,3,'TRUE','Mumbai, Pune, Bangalore, Hyderabad','Analytics','Pure analytics firm; BFSI clients; fresher friendly'),
('Mphasis','https://careers.mphasis.com/','mphasis',1,3,'TRUE','Bangalore, Hyderabad, Pune','Analytics/IT','Banking-focused analytics'),
('CRISIL','https://www.crisil.com/en/home/careers.html','crisil',1,3,'TRUE','Mumbai, Pune, Bangalore','Credit Analytics','S&P Global subsidiary; Global Analytics Centre 3000+ analysts'),
# TIER 2
('BlackRock','https://careers.blackrock.com/','blackrock',2,3,'TRUE','Mumbai','Asset Management','Aladdin platform; one of best DS employers in finance globally'),
('AlphaSense','https://www.alpha-sense.com/about/careers/','alphasense',2,3,'TRUE','Mumbai, Bangalore','Financial AI','RAG over financial documents is their core product; perfect fit'),
('Kensho (S&P Global)','https://careers.spglobal.com/','kensho-technologies',2,3,'TRUE','Hyderabad','Financial AI','S&P Global AI subsidiary; NLP + ML for market data'),
('Bloomberg','https://careers.bloomberg.com/','bloomberg-lp',2,3,'TRUE','Mumbai','Market Data','Financial NLP + AI for terminal features'),
('LSEG (Refinitiv)','https://careers.lseg.com/','london-stock-exchange-group',2,3,'TRUE','Bangalore, Hyderabad','Market Data','AI/ML for market data; large India team'),
('Quantiphi','https://quantiphi.com/careers/','quantiphi',2,3,'TRUE','Mumbai, Bangalore','AI/ML Consulting','AI/ML consulting; strong BFSI vertical; RAG + agents'),
('Goldman Sachs','https://www.goldmansachs.com/careers/','goldman-sachs',2,3,'TRUE','Bangalore, Hyderabad','Investment Bank GCC','Selective but analyst programs exist'),
('Deutsche Bank','https://careers.db.com/','deutsche-bank',2,3,'TRUE','Pune, Bangalore, Mumbai','Investment Bank GCC','Large analytics GCC; data engineering + DS'),
('Barclays','https://home.barclays/careers/','barclays',2,3,'TRUE','Pune','Investment Bank GCC','Pune GCC hires analytics regularly'),
('HSBC','https://www.hsbc.com/careers','hsbc',2,3,'TRUE','Hyderabad, Pune, Bangalore','Banking GCC','Large analytics hub'),
('Morgan Stanley','https://www.morganstanley.com/people-opportunities','morgan-stanley',2,2,'TRUE','Mumbai, Bangalore','Investment Bank','Selective analyst programs'),
('BNY Mellon','https://www.bnymellon.com/us/en/careers.html','bny-mellon',2,2,'TRUE','Pune','Investment Bank GCC','Analytics GCC; data roles'),
('State Street','https://careers.statestreet.com/','state-street',2,2,'TRUE','Bangalore, Hyderabad','Investment Bank GCC','Analytics roles in Bangalore GCC'),
('American Express','https://jobs.americanexpress.com/','american-express',2,3,'TRUE','Bangalore','Financial Services GCC','Strong data science team; India GCC'),
('Mastercard','https://careers.mastercard.com/','mastercard',2,3,'TRUE','Pune, Bangalore','Financial Services GCC','Data + AI team; structured hiring'),
('Visa','https://careers.visa.com/','visa',2,3,'TRUE','Bangalore, Mumbai','Financial Services GCC','Data science roles in Bangalore GCC'),
('Fidelity Investments','https://jobs.fidelity.com/','fidelity-investments',2,3,'TRUE','Bangalore','Asset Management GCC','Investment analytics GCC'),
('Synchrony Financial','https://www.synchronycareers.com/','synchrony',2,2,'TRUE','Hyderabad','Financial Services GCC','Analytics GCC; credit risk data'),
('Capital One','https://www.capitalonecareers.com/','capital-one',2,2,'TRUE','Bangalore','Banking GCC','Tech-first bank; strong data culture'),
('S&P Global','https://careers.spglobal.com/','s-p-global',2,2,'TRUE','Hyderabad, Pune','Market Data','Analytics roles; market data'),
("Moody's Analytics",'https://careers.moodys.com/','moodys',2,2,'TRUE','Mumbai','Market Data','Risk analytics roles'),
('FactSet','https://factset.wd1.myworkdayjobs.com/FactSetCareers','factset',2,2,'TRUE','Hyderabad, Mumbai','Market Data','Financial data analytics roles'),
('TransUnion','https://careers.transunion.com/','transunion',2,3,'TRUE','Pune, Hyderabad','Credit Analytics','Credit analytics; data science'),
('Razorpay','https://razorpay.com/jobs/','razorpay',2,3,'TRUE','Bangalore','Fintech','Scale fintech; strong data team'),
('PhonePe','https://www.phonepe.com/en/careers.html','phonepe',2,3,'TRUE','Bangalore','Fintech','Large fintech analytics team'),
('CRED','https://careers.cred.club/','dreamplug-technologies',2,2,'TRUE','Bangalore','Fintech','Data-first fintech'),
('Groww','https://groww.in/open-positions','groww',2,2,'TRUE','Bangalore','Fintech','Growing fintech; data roles'),
('Paytm','https://paytm.com/about-us/careers/','paytm',2,2,'TRUE','Bangalore','Fintech','Large data team'),
('Bajaj Finance','https://www.bajajfinserv.in/careers','bajaj-finserv',2,3,'TRUE','Pune, Mumbai, Bangalore','NBFC','Largest NBFC; credit risk + analytics'),
('Kotak Mahindra Bank','https://www.kotak.com/en/careers.html','kotak-mahindra-bank',2,2,'TRUE','Mumbai, Pune','Banking','Growing analytics division'),
('Sigmoid Analytics','https://sigmoid.com/careers/','sigmoid-analytics',2,2,'TRUE','Bangalore, Hyderabad','Analytics','Analytics firm; BFSI focused'),
('EClerx','https://www.eclerx.com/careers/','eclerx',2,2,'TRUE','Mumbai, Pune, Bangalore','Analytics BPO','Financial services BPO; analytics'),
('Nexdigm','https://www.nexdigm.com/careers/','nexdigm',2,2,'TRUE','Mumbai, Pune','Finance Consulting','Finance tax + risk advisory; analytics roles'),
('BCG','https://careers.bcg.com/','boston-consulting-group',2,2,'TRUE','Mumbai, Bangalore','Consulting','Data analytics practice; selective'),
('Aon','https://careers.aon.com/','aon',2,2,'TRUE','Mumbai, Bangalore, Pune','Insurance/Risk','Risk + insurance analytics'),
('Morningstar','https://www.morningstar.com/careers','morningstar',2,2,'TRUE','Mumbai, Hyderabad','Investment Research','Investment research analytics; AI on fund data'),
('Dezerv','https://dezerv.in/careers','dezerv',2,2,'TRUE','Bangalore','Wealth Management','AI-driven portfolio management fintech'),
('Smallcase','https://smallcase.com/about/careers','smallcase',2,2,'TRUE','Bangalore','WealthTech','Curated investment portfolios; strong data/ML team'),
('INDmoney','https://indmoney.com/careers','indmoney',2,2,'TRUE','Bangalore','WealthTech','Super finance app; strong data team'),
('Jarvis Invest','https://www.jarvisinvest.com/careers','jarvis-invest',2,2,'TRUE','Bangalore','AI WealthTech','AI-driven portfolio management; uses ML/AI at core'),
("ICRA (Moody's)",'https://www.icra.in/Careers','icra',2,2,'TRUE','Mumbai, Gurgaon','Credit Analytics',"Moody's subsidiary; credit ratings + research analytics"),
('CARE Ratings','https://www.careratings.com/Career.aspx','care-ratings',2,2,'TRUE','Mumbai','Credit Analytics','Credit rating + financial research analytics'),
('Dun & Bradstreet India','https://www.dnb.com/about-us/careers.html','dun-bradstreet',2,2,'TRUE','Mumbai','Business Analytics','Business credit data + risk analytics'),
('Karza Technologies','https://karza.in/careers.html','karza-technologies',2,2,'TRUE','Mumbai','FinData','Financial data APIs for banks/NBFCs; fraud + KYC analytics'),
('Bureau','https://bureau.id/careers','bureau-id',2,2,'TRUE','Bangalore','Risk Analytics','Fraud + risk analytics for financial services'),
('Signzy','https://signzy.com/careers/','signzy',2,2,'TRUE','Bangalore','AI Banking','AI for banking KYC + onboarding; NLP/RAG relevant'),
('Yubi (CredAvenue)','https://go.yubi.in/careers','yubi',2,2,'TRUE','Chennai, Bangalore','Debt Analytics','Debt platform + credit analytics'),
('Capco','https://www.capco.com/careers','capco',2,2,'TRUE','Mumbai, Bangalore','FinTech Consulting','Financial technology consulting; data + AI for banks'),
('ZS Associates','https://www.zs.com/careers','zs-associates',2,2,'TRUE','Pune, Bangalore','Analytics Consulting','Analytics consulting; growing BFSI AI practice'),
('Franklin Templeton India','https://www.franklintempleton.co.in/investor/about-us/career-center','franklin-templeton',2,2,'TRUE','Mumbai','Asset Management','Portfolio analytics; investment data'),
('Clearwater Analytics','https://clearwateranalytics.com/careers/','clearwater-analytics',2,2,'TRUE','Pune','Investment Analytics','Investment accounting + analytics platform'),
('TCS (BFSI)','https://ibegin.tcs.com/iBegin/','tata-consultancy-services',2,2,'TRUE','Mumbai, Pune, Bangalore, Hyderabad','IT/Analytics','Largest BFSI IT employer; BFSI analytics division'),
('Infosys BPM','https://careers.infosysbpm.com/','infosys-bpm',2,2,'TRUE','Bangalore, Hyderabad, Pune','Analytics BPO','BFSI analytics + AI practice'),
('Tech Mahindra (BFSI)','https://careers.techmahindra.com/','tech-mahindra',2,2,'TRUE','Pune, Bangalore, Hyderabad','IT/Analytics','BFSI tech + analytics'),
('MassMutual India','https://jobs.massmutual.com/','massmutual',2,2,'TRUE','Hyderabad','Insurance GCC','Insurance + investment analytics GCC'),
# TIER 3
('UBS','https://www.ubs.com/global/en/careers.html','ubs',3,1,'TRUE','Hyderabad','Investment Bank GCC','Small India GCC for analytics'),
('Nomura','https://www.nomura.com/careers/','nomura',3,1,'TRUE','Mumbai, Bangalore','Investment Bank','Selective; quant-focused'),
('BNP Paribas','https://group.bnpparibas/en/careers','bnp-paribas',3,1,'TRUE','Mumbai, Bangalore','Investment Bank','Limited fresher data roles'),
('Societe Generale','https://careers.societegenerale.com/','societe-generale',3,1,'TRUE','Bangalore','Investment Bank GCC','Data engineering heavy'),
('Northern Trust','https://www.northerntrust.com/united-states/what-we-do/about-us/careers','northern-trust',3,1,'TRUE','Bangalore','Banking GCC','Small India analytics team'),
('Macquarie','https://www.macquarie.com/au/en/careers.html','macquarie',3,1,'TRUE','Mumbai, Hyderabad','Investment Bank','Selective; finance analytics'),
('Standard Chartered','https://www.sc.com/en/careers/','standard-chartered-bank',3,1,'TRUE','Bangalore, Mumbai','Banking','Moderate data hiring'),
('MSCI','https://www.msci.com/careers','msci',3,1,'TRUE','Mumbai, Hyderabad','Market Data','Niche market risk analytics'),
('Nasdaq','https://www.nasdaq.com/about/careers','nasdaq',3,1,'TRUE','Mumbai, Bangalore','Market Data','Limited India DS roles'),
('SS&C Technologies','https://www.ssctech.com/company/careers','ss-c-technologies',3,1,'TRUE','Bangalore, Hyderabad','FinTech GCC','Fund accounting tech'),
('Broadridge Financial','https://careers.broadridge.com/','broadridge-financial-solutions',3,1,'TRUE','Hyderabad, Mumbai','FinTech GCC','Limited DS roles'),
('FIS Global','https://careers.fisglobal.com/','fis',3,1,'TRUE','Pune, Bangalore, Hyderabad','FinTech GCC','Tech-heavy; limited DS freshers'),
('Fiserv','https://careers.fiserv.com/','fiserv',3,1,'TRUE','Pune, Bangalore, Hyderabad','FinTech GCC','Same as FIS'),
('HDFC Life','https://www.hdfclife.com/about-us/careers','hdfc-life',3,1,'TRUE','Mumbai, Bangalore','Insurance','Insurance analytics'),
('ICICI Prudential Life','https://www.iciciprulife.com/about-us/careers.html','icici-prudential-life-insurance-company-limited',3,1,'TRUE','Mumbai, Pune','Insurance','Insurance analytics'),
('Bajaj Allianz','https://www.bajajallianz.com/corp/about-us/careers.html','bajaj-allianz-general-insurance',3,1,'TRUE','Pune','Insurance','Insurance data roles'),
('Acko','https://www.acko.com/careers/','acko',3,1,'TRUE','Bangalore','InsurTech','Growing data team'),
('Digit Insurance','https://www.godigit.com/careers','go-digit-general-insurance',3,1,'TRUE','Bangalore','InsurTech','Growing; data roles emerging'),
('Zerodha','https://zerodha.com/careers/','zerodha',3,1,'TRUE','Bangalore','Fintech','Small team; selective'),
('Upstox','https://upstox.com/open-positions/','upstox',3,1,'TRUE','Mumbai, Bangalore','Fintech','Growing; data roles'),
('Angel One','https://www.angelone.in/careers','angel-broking',3,1,'TRUE','Mumbai','Fintech','Analytics team; moderate hiring'),
('BharatPe','https://bharatpe.com/careers','bharatpe',3,1,'TRUE','Bangalore','Fintech','Fintech analytics roles'),
('Fibe','https://fibe.in/careers/','fibe',3,1,'TRUE','Bangalore, Mumbai','Fintech','Credit analytics; growing'),
('KreditBee','https://www.kreditbee.in/careers','kreditbee',3,1,'TRUE','Bangalore','Fintech','Credit risk DS roles'),
('Wipro (BFSI)','https://careers.wipro.com/','wipro',3,1,'TRUE','Bangalore, Hyderabad, Pune','IT','BFSI practice; limited DS freshers'),
('HCLtech (BFSI)','https://www.hcltech.com/careers','hcltech',3,1,'TRUE','Bangalore, Hyderabad','IT','BFSI analytics; limited freshers'),
('Nielsen','https://careers.nielsen.com/','nielsen',3,1,'TRUE','Mumbai, Bangalore','Analytics','Market analytics; limited pure finance'),
('Tracxn','https://tracxn.com/careers','tracxn',3,1,'TRUE','Bangalore','Investment Intelligence','Deal discovery for VC/PE; small data team'),
('India Ratings (Fitch)','https://www.indiaratings.co.in/career','india-ratings-and-research',3,1,'TRUE','Mumbai','Credit Analytics','Fitch subsidiary; credit analytics'),
('KKR','https://www.kkr.com/careers','kkr',3,1,'TRUE','Gurgaon','PE/VC','Portfolio analytics; rare but exists'),
('Blackstone','https://www.blackstone.com/careers/','blackstone',3,1,'TRUE','Mumbai','PE/VC','Portfolio analytics; very selective'),
('Fisdom','https://www.fisdom.com/careers/','fisdom',3,1,'TRUE','Bangalore','WealthTech','Wealth management platform'),
('ET Money','https://www.etmoney.com/careers','et-money',3,1,'TRUE','Gurgaon','WealthTech','Personal finance app; analytics'),
('Credgenics','https://credgenics.com/careers/','credgenics',3,1,'TRUE','Bangalore','AI Collections','AI for loan collections + risk analytics'),
('CIBC India','https://careers.cibc.com/','cibc',3,1,'TRUE','Bangalore','Banking GCC','Canadian bank GCC; analytics roles'),
]

for i, row in enumerate(companies, 2):
    ws1.append(list(row))
    row_style(ws1, i, i % 2 == 0)
    tier_cell(ws1, i, 4, row[3])

# ── SHEET 2: ROLES ────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Roles')
ws2.freeze_panes = 'A2'
header(ws2,
    ['Role Title','Search Keywords (comma separated)','Tags','Finance Application','Active'],
    [28,62,38,44,8], '1A5276')

roles = [
('Data Scientist','data scientist, DS, machine learning, ML, statistical modelling, predictive modelling, data science','ml, python, statistics, sklearn, r, sql','Credit risk modelling, fraud detection, portfolio analytics, churn prediction','TRUE'),
('Data Analyst','data analyst, business analyst, analytics, BI analyst, reporting analyst, insights analyst, financial analyst data','sql, python, excel, tableau, powerbi, analytics','Financial reporting, trading analytics, investment performance reporting','TRUE'),
('AI Engineer','AI engineer, LLM engineer, generative AI engineer, NLP engineer, applied AI, AI developer, GenAI','llm, genai, langchain, python, nlp, transformers, rag','Financial document AI, earnings call analysis, RAG over filings','TRUE'),
('BI Developer','BI developer, BI engineer, Power BI developer, Tableau developer, business intelligence, dashboard developer','powerbi, tableau, sql, bi, dashboard, dax','Financial dashboards, KPI reporting, investment performance visualization','TRUE'),
('Machine Learning Engineer','ML engineer, machine learning engineer, MLOps engineer, AI ML engineer','ml, python, sklearn, mlops, tensorflow, pytorch','Fraud detection models, credit scoring, algorithmic trading signals','TRUE'),
('Data Engineer','data engineer, DE, data pipeline, ETL engineer, data platform engineer, big data engineer','python, sql, spark, kafka, airflow, dbt, data pipeline','Financial data pipelines, market data ingestion, trade data processing','TRUE'),
('Investment Banking Analytics','investment banking analytics, IB analytics, IBD analyst, deal analytics, M&A analyst data, capital markets analyst','finance, excel, python, sql, valuation, modelling','Deal support analytics, market research, financial modelling for IB','TRUE'),
('Financial Data Analyst','financial data analyst, finance analyst, FP&A analyst, financial analyst, financial reporting analyst','finance, sql, excel, python, powerbi, tableau','P&L analysis, budgeting, financial planning analytics','TRUE'),
('Quantitative Analyst','quant analyst, quantitative analyst, quantitative researcher, quant developer, quant finance','python, r, statistics, finance, derivatives, risk, c++','Risk modelling, derivatives pricing, portfolio optimization','TRUE'),
('Risk Analyst','risk analyst, credit risk analyst, market risk analyst, risk modelling, risk management analyst','risk, statistics, python, sql, finance, credit, basel','Credit risk, market risk, operational risk, regulatory analytics','TRUE'),
('NLP Engineer','NLP engineer, natural language processing, text analytics, computational linguistics, NLP data scientist','nlp, python, transformers, huggingface, spacy, bert, llm','Financial news analytics, earnings call NLP, regulatory text mining','TRUE'),
('Research Analyst','research analyst, equity research analyst, investment research, financial research, market research analyst','finance, excel, python, sql, research, equity','Equity research, market intelligence, investment thesis analytics','TRUE'),
]

for i, row in enumerate(roles, 2):
    ws2.append(list(row))
    row_style(ws2, i, i % 2 == 0)

# ── SHEET 3: HR CONTACTS ─────────────────────────────────────────────────────
ws3 = wb.create_sheet('HR Contacts')
ws3.freeze_panes = 'A2'
header(ws3,
    ['Name','Company','Email','Sector','Tier','Contacted','Date Contacted','Response','Notes'],
    [24,22,36,20,6,12,16,16,28], '145A32')
ws3.auto_filter.ref = 'A1:I1'

contacts = [
('Shantanu Verma','EXL','shantanu.verma@exlservice.com','Analytics BPO',1,'FALSE','','',''),
('Pragati Agrawal','EXL','pragati.agrawal@exlservice.com','Analytics BPO',1,'FALSE','','',''),
('Vibhuti Gupta','EXL','vibhuti.gupta1@exlservice.com','Analytics BPO',1,'FALSE','','',''),
('Hitesh Goyal','EXL','hitesh.goyal@exlservice.com','Analytics BPO',1,'FALSE','','',''),
('Nandini Sharma','EXL','nandini.sharma@exlservice.com','Analytics BPO',1,'FALSE','','',''),
('Mohan Meher','Genpact','mohan.meher@genpact.com','Analytics BPO',1,'FALSE','','',''),
('Akankasha Talwar','Genpact','akankasha.talwar@genpact.com','Analytics BPO',1,'FALSE','','',''),
('Rajni Bala','Genpact','rajni.bala@genpact.com','Analytics BPO',1,'FALSE','','',''),
('Bidisha Banerjee','Genpact','bidisha.banerjee1@genpact.com','Analytics BPO',1,'FALSE','','',''),
('Manvi Bhutani','Genpact','manvi.bhutani1@genpact.com','Analytics BPO',1,'FALSE','','',''),
('Lochan Dixit','Genpact','lochan.dixit@genpact.com','Analytics BPO',1,'FALSE','','',''),
('Tanurima Mukherjee','Genpact','tanurima.mukherjee@genpact.com','Analytics BPO',1,'FALSE','','',''),
('Havisha Arora','Genpact','havisha.arora@genpact.com','Analytics BPO',1,'FALSE','','',''),
('Sharutika Sharutika','Genpact','Sharutika.Sharutika@genpact.com','Analytics BPO',1,'FALSE','','',''),
('Ashwini Jagtap','WNS','Ashwini.jagtap@wns.com','Analytics BPO',1,'FALSE','','',''),
('Anuj Verma','Accenture','Anuj.d.verma@accenture.com','Consulting',1,'FALSE','','',''),
('Tanya Jain','Accenture','tanya.h.jain@accenture.com','Consulting',1,'FALSE','','',''),
('Soma Pramanik','Capgemini','soma.pramanik@capgemini.com','Consulting',1,'FALSE','','',''),
('Monica Ashok','EY','monica.ashok@in.ey.com','Big 4',1,'FALSE','','',''),
('Shipra Mishra','PwC','Shipra.mishra@pwc.com','Big 4',1,'FALSE','','',''),
('Swati Bewal','KPMG','swatibewal@kpmg.com','Big 4',1,'FALSE','','',''),
('Rohit Naidu','KPMG','rohitnaidu@kpmg.com','Big 4',1,'FALSE','','',''),
('Varun Chawla','KPMG','varunchawla@kpmg.com','Big 4',1,'FALSE','','',''),
('Ananya Srivastava','KPMG','ananyasrivastava@kpmg.com','Big 4',1,'FALSE','','',''),
('Preeti Chauhan','KPMG','preetichauhan@kpmg.com','Big 4',1,'FALSE','','',''),
('Moni Kachoudhary','KPMG','on-monikachoudhary@kpmg.com','Big 4',1,'FALSE','','',''),
('Pbhargava','Deloitte','pbhargava@deloitte.com','Big 4',1,'FALSE','','',''),
('Sneha Mishra','HDFC Bank','sneha.mishra@hdfclife.com','Banking',1,'FALSE','','',''),
('Ashik Sarsiha','Citi','ashik.sarsiha@citi.com','Banking GCC',1,'FALSE','','',''),
('Nandita Chaudhary','Citi','nandita.chaudhary@citi.com','Banking GCC',1,'FALSE','','',''),
('Saikat Banerjee','HSBC','saikat.banerjee@hsbc.co.in','Banking GCC',2,'FALSE','','',''),
('Amitkumar Das','HSBC','amitkumardas@hsbc.co.in','Banking GCC',2,'FALSE','','',''),
('Ganesh Kumar','Barclays','ganesh.kumar4@barclays.com','Investment Bank GCC',2,'FALSE','','',''),
('Ishant Gupta','IndusInd Bank','Ishant.gupta@indusind.com','Banking',1,'FALSE','','',''),
('Sweta Sharma','Kotak','sweta.sharma2@kotak.com','Banking',2,'FALSE','','',''),
('Sushma Vedula','BCG','vedula.sushma@bcg.com','Consulting',2,'FALSE','','',''),
('Shashwathi P','BCG','P.Shashwathi@bcg.com','Consulting',2,'FALSE','','',''),
('Priyanka Toraskar','TransUnion','priyanka.toraskar@transunion.com','Credit Analytics',2,'FALSE','','',''),
('Ridhi Garg','Aon','ridhi.garg3@aon.com','Insurance/Risk',2,'FALSE','','',''),
('Ashish Sharma','American Express','ashish.sharma1@aexp.com','Financial Services GCC',2,'FALSE','','',''),
('Shashank Chauhan','American Express','shashank.chauhan@aexp.com','Financial Services GCC',2,'FALSE','','',''),
('Kanika Sharma','American Express','kanika.sharma17@aexp.com','Financial Services GCC',2,'FALSE','','',''),
('Seyyid M Hussain','American Express','seyyid.m.hussain@aexp.com','Financial Services GCC',2,'FALSE','','',''),
('Adarsh Purwar','Paytm','adarsh.purwar@paytm.com','Fintech',2,'FALSE','','',''),
('Nidhi Chawda','Razorpay','nidhi.chawda@razorpay.com','Fintech',2,'FALSE','','',''),
('Somya Bagaria','Cashflo','somya.bagaria@cashflo.io','Fintech',2,'FALSE','','',''),
('Siddharoodha Biradar','Fibe','siddharoodha.biradar@fibe.in','Fintech',3,'FALSE','','',''),
('Ritika Mahawer','India Lends','ritika.mahawer@indialends.com','Fintech',3,'FALSE','','',''),
('Nimisha Singh','Nexdigm','Nimisha.singh@nexdigm.com','Finance Consulting',2,'FALSE','','',''),
('Siddhi Narvankar','Nexdigm','siddhi.narvankar@nexdigm.com','Finance Consulting',2,'FALSE','','',''),
('Archana Kumari','Nexdigm','archana.kumari@nexdigm.com','Finance Consulting',2,'FALSE','','',''),
('Sukanya K','Nielsen','sukanya.k2@nielsen.com','Analytics',3,'FALSE','','',''),
('Marpita','Sigmoid Analytics','marpita@sigmoidanalytics.com','Analytics',2,'FALSE','','',''),
('Vipul Jha','EClerx','vipul.jha@eclerx.com','Analytics BPO',2,'FALSE','','',''),
('Shivani Srivastava','Xceedance','shivani.srivastava@xceedance.com','Insurance Analytics',3,'FALSE','','',''),
('HR Team','PhonePe','analytics-careers@phonepe.com','Fintech',2,'FALSE','','',''),
]

for i, row in enumerate(contacts, 2):
    ws3.append(list(row))
    row_style(ws3, i, i % 2 == 0)
    tier_cell(ws3, i, 5, row[4])

out = 'C:/Users/dheem/Documents/career-scout/Career_Scout_Master.xlsx'
wb.save(out)
print(f'Saved: {out}')
print(f'Companies : {len(companies)}')
print(f'Roles     : {len(roles)}')
print(f'HR Contacts: {len(contacts)}')
